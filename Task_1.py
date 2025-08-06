import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from collections import defaultdict


class ExcelFilter:
    def __init__(self, master):
        self.master = master
        self.master.title("")
        self.master.geometry("400x400+100+100")

        self.excel_path_var = tk.StringVar()
        self.new_excel_path_var = tk.StringVar()
        self.var_column = tk.StringVar()
        self.var = tk.StringVar()

        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.create_widgets()

    def create_widgets(self):
        tk.Label(self.master, text="Выберите Excel файл:").pack(pady=5)
        tk.Entry(self.master, textvariable=self.excel_path_var, width=50).pack(pady=5)
        tk.Button(self.master, text="Обзор", command=self.select_excel_file).pack(pady=5)

        tk.Label(self.master, text="Сохранить как новый Excel файл:").pack(pady=5)
        tk.Entry(self.master, textvariable=self.new_excel_path_var, width=50).pack(pady=5)
        tk.Button(self.master, text="Обзор", command=self.select_new_excel_file).pack(pady=5)

        tk.Label(self.master, text="Введите значение столбца").pack(pady=5)
        tk.Entry(self.master, textvariable=self.var_column, width=50).pack(pady=5)
        tk.Label(self.master, text="Введите значение").pack(pady=5)
        tk.Entry(self.master, textvariable=self.var, width=50).pack(pady=5)

        tk.Button(self.master, text="Преобразовать", command=self.filtered_excel).pack(pady=20)

    def select_excel_file(self):
        pdf_file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        self.excel_path_var.set(pdf_file)

    def select_new_excel_file(self):
        excel_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        self.new_excel_path_var.set(excel_file)

    def filtered_excel(self):
        excel_path = self.excel_path_var.get()
        new_excel_path = self.new_excel_path_var.get()
        var_column = self.var_column.get().strip().lower()
        var = self.var.get().strip().lower()
        result_dict = defaultdict(list)

        if not excel_path or not new_excel_path:
            messagebox.showerror("Ошибка", "Пожалуйста, выберите файлы.")
            return

        if not var_column or not var:
            messagebox.showerror("Ошибка", "Пожалуйста, введите значения.")
            return

        file = openpyxl.load_workbook(excel_path)
        sheet = file.active
        headers = [cell.value for cell in sheet[9]]
        columns = ["фио", "должность", "отдел", "дата найма", "зарплата"]

        try:
            col_index = columns.index(var_column)
        except ValueError:
            messagebox.showerror("Ошибка", "Пожалуйста, введите существующее значение столбца.")
            return

        for row in sheet.iter_rows(min_row=10, values_only=True):
            if row[0]:
                result_dict["ФИО"].append(row[1])
                result_dict["Должность"].append(row[4])
                result_dict["Отдел"].append(row[6])
                result_dict["Дата найма"].append(row[7])
                result_dict["Зарплата"].append(row[5])

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(columns)

        for i in range(len(headers)):
            if result_dict[list(result_dict.keys())[col_index]][i].lower() == var:
                row = [result_dict[key][i] for key in result_dict.keys()]
                worksheet.append(row)

        workbook.save(new_excel_path)
        messagebox.showinfo("Успех", f"Файл успешно сохранен: {new_excel_path}")


if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelFilter(root)
    root.mainloop()
